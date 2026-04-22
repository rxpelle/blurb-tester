"""Historical KENP backfill from KDP's native XLSX report.

The atAGlance endpoint is capped at a 22-day rolling window. To go further
back (months or years), KDP offers an XLSX download from the Reports UI —
but the JSON initiation endpoint (POST /download/report/kenpread/{locale}/…)
requires a request-body shape that is stitched together from Redux UI state
and varies per account. Fully reverse-engineering it would be brittle.

Pragmatic path: the user downloads the XLSX once from the KDP UI, then this
module parses it and upserts into ku_page_flip's DB.

Expected XLSX layout (observed from KDP "KENP Read" report, 2026 format):
  Sheet per marketplace OR single "Combined" sheet
  Columns (may vary): Date, Title, ASIN, Marketplace, KENP Read
"""

import re
from datetime import date as _date, datetime
from pathlib import Path

from openpyxl import load_workbook


_DATE_CANDIDATES = ("date", "order date")
_ASIN_CANDIDATES = ("asin",)
_TITLE_CANDIDATES = ("title", "title name", "book title")
_MARKETPLACE_CANDIDATES = ("marketplace", "territory")
_KENP_CANDIDATES = ("kenp read", "kenpc", "pages read", "kenp", "kindle edition normalized pages")


def _find_col(headers: list[str], candidates: tuple[str, ...]) -> int:
    norm = [(h or "").strip().lower() for h in headers]
    for c in candidates:
        if c in norm:
            return norm.index(c)
    for c in candidates:
        for i, h in enumerate(norm):
            if c in h:
                return i
    return -1


def _parse_date(val) -> _date | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, _date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d %b %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return _date(int(m[1]), int(m[2]), int(m[3]))
    return None


def _parse_int(val) -> int:
    if val is None or val == "":
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    try:
        return int(str(val).replace(",", "").strip())
    except ValueError:
        return 0


def parse_kenp_xlsx(xlsx_path: Path) -> list[dict]:
    """Return rows: {asin, title, date, kenp_reads, marketplace}.

    Walks every sheet, finds header row by looking for a 'date' column,
    then emits one row per non-zero data row.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    rows: list[dict] = []
    for sheet in wb.worksheets:
        data_iter = sheet.iter_rows(values_only=True)
        header_row = None
        header_idx = None
        for idx, r in enumerate(data_iter):
            if r and any(
                isinstance(c, str) and c.strip().lower() in _DATE_CANDIDATES
                for c in r
            ):
                header_row = list(r)
                header_idx = idx
                break
        if not header_row:
            continue

        date_col = _find_col(header_row, _DATE_CANDIDATES)
        asin_col = _find_col(header_row, _ASIN_CANDIDATES)
        title_col = _find_col(header_row, _TITLE_CANDIDATES)
        mp_col = _find_col(header_row, _MARKETPLACE_CANDIDATES)
        kenp_col = _find_col(header_row, _KENP_CANDIDATES)
        if date_col < 0 or kenp_col < 0:
            continue

        sheet_mp = sheet.title if mp_col < 0 else None

        for r in data_iter:
            if not r:
                continue
            dt = _parse_date(r[date_col] if date_col < len(r) else None)
            reads = _parse_int(r[kenp_col] if kenp_col < len(r) else None)
            if not dt or reads <= 0:
                continue
            asin = (str(r[asin_col]).strip() if asin_col >= 0 and asin_col < len(r) and r[asin_col] else "")
            title = (str(r[title_col]).strip() if title_col >= 0 and title_col < len(r) and r[title_col] else "")
            marketplace = (
                str(r[mp_col]).strip() if mp_col >= 0 and mp_col < len(r) and r[mp_col]
                else sheet_mp or "US"
            )
            rows.append({
                "asin": asin,
                "title": title,
                "date": dt.isoformat(),
                "kenp_reads": reads,
                "marketplace": marketplace,
            })
    wb.close()
    rows.sort(key=lambda r: (r["asin"], r["date"], r["marketplace"]))
    return rows


def import_xlsx_to_kupf(xlsx_path: Path) -> dict:
    """Parse XLSX and upsert into ku_page_flip DB. Idempotent."""
    from ku_page_flip.db import Database

    rows = parse_kenp_xlsx(xlsx_path)
    if not rows:
        return {"rows": 0, "books": {}}

    by_asin: dict[str, list[dict]] = {}
    titles: dict[str, str] = {}
    for r in rows:
        if not r["asin"]:
            continue
        by_asin.setdefault(r["asin"], []).append(r)
        if r["title"]:
            titles[r["asin"]] = r["title"]

    per_book: dict[str, int] = {}
    with Database() as db:
        for asin, asin_rows in by_asin.items():
            book = db.get_book_by_asin(asin) or db.add_book(
                title=titles.get(asin, asin), asin=asin
            )
            for r in asin_rows:
                db.add_daily_read(
                    book_id=book.id,
                    dt=_date.fromisoformat(r["date"]),
                    kenp_reads=r["kenp_reads"],
                    marketplace=r["marketplace"],
                )
            per_book[asin] = len(asin_rows)
    return {"rows": len(rows), "books": per_book, "titles": titles}
